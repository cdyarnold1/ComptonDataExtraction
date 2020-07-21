import json
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import math

SURE_REPLACEMENTS = {
    "€": "E",
    "S": "5",
    "F": "E",
    ",": ".",
    "O": "0"
}

# will get marked orange
POSSIBLE_REPLACEMENTS = {
    "B": "8",
    "A": "8",
    "R": "8",
    "I": "1",
    "l": "1",
    "T": "1"
}

def nth_repl(s, sub, repl, n):
    find = s.find(sub)
    # If find is not -1 we have found at least one match for the substring
    i = find != -1
    # loop util we find the nth or we find no match
    while find != -1 and i != n:
        # find + 1 means we start searching from after the last match
        find = s.find(sub, find + 1)
        i += 1
    # If i is equal to n we found nth match so replace
    if find < 0:
        return s
    if i == n:
        return s[:find] + repl + s[find+len(sub):]
    return s

# Open output from text detection
with open('p33_h1_silver.json', 'r') as f:
    data = json.load(f)

# data is a list of words (which may or may not be a full number) represented as tuple: ((xbound1, xbound2), (ybound1, ybound2), text)
print(data)

# Calculate where the rows are
upper_bounds = []
lower_bounds = []

for word in data:
    upper_bounds.append(min(word[1]))
    lower_bounds.append(max(word[1]))

upper_bounds = np.histogram(upper_bounds, bins=list(range(0, max(upper_bounds)+10, 3)))[0]
lower_bounds = np.histogram(lower_bounds, bins=list(range(0, max(lower_bounds)+10, 3)))[0]

# Row upper bounds
lines_up = find_peaks(upper_bounds, distance=10, height=2)[0]
# Row lower bounds
lines_down = find_peaks(lower_bounds, distance=10, height=2)[0]

# Sometimes the upper bound doesn't get detected. So add it
if len(lines_up) < len(lines_down):
    lines_up = np.insert(lines_up, 0, 0)

# Show a graph of the row detection
plt.figure("Detected rows")
plt.plot(upper_bounds)
plt.plot(lower_bounds)
plt.scatter(lines_up, np.zeros(len(lines_up)))
plt.scatter(lines_down, np.zeros(len(lines_down)))
plt.xlabel("pixels/3")
plt.show()

# Detect where the columns are
cols = []
maxX = max(data, key=lambda word: max(word[0]))
cols = np.zeros(max(maxX[0])+20)
for word in data:
    for i in range(*word[0]):
        cols[i] += 1
cols_inv = -cols

cols_inv = np.append(cols_inv,-100)
cols_inv = np.insert(cols_inv,0, -100)

# X values between columns
seps = find_peaks(cols_inv, distance=10, plateau_size=10, height=0)[0]

# Show the detected columns
plt.plot(cols)
plt.plot(cols_inv)
plt.scatter(seps, np.zeros(len(seps)))
plt.show()

# Some rows have two decimal places, some have only one
decimals = np.full(len(lines_up), 0)
decimals[0:25] = 2
decimals[25:] = 1

# Excel table rows
lines = []
# Cells that should be orange
attentions = []
# Cells that should be red
warns = []
i = 0

# Iterate through the detected rows
for line in zip(sorted(lines_up), sorted(lines_down)):
    # Midpoint of the row
    avg_height = sum(line)*3/2
    values = [""]*(len(seps))
    # rightmost x value of last word
    last_end = 1000000

    # Go through all words and check if they are on the avg_height line
    for word in data:
        if min(word[1]) <= avg_height <= max(word[1]):
            # Get the index of the column the word belongs to
            idx = np.argmax(seps>sum(word[0])/2)-1
            # Get the text
            t = word[2]

            if i == 0: # If its a header
                values[idx] += t
                last_end = max(word[0])
                continue

            # Replace common mismatches
            for w, r in SURE_REPLACEMENTS.items():
                t = t.replace(w, r)
            for w, r in POSSIBLE_REPLACEMENTS.items():
                if w in t:
                    attentions.append((idx, i))
                t = t.replace(w, r)
            # If it is the block of text of a number, it has to be the decimal
            if len(values[idx]) == 0:
                try:
                    # Round it to the defined amount of decimal places. That gets rid of misinterpretations of the 'E'
                    n = str(math.floor(float(t.split("E")[0])*(10**decimals[i]))/(10.0**decimals[i]))
                    # Reassemble the text
                    r = t.split("E")[1] if len(t.split("E")) > 1 else ""
                    t = n+"E"+r
                except:
                    pass
            # Otherwise it's the exponent
            else:
                t = t.replace(".", "") # This cannot have a '.' (result of misreading the '+' sign)
            values[idx] += t
            last_end = max(word[0])

    if i > 0:
        # Do some more fixing now that we have the full number
        for j, val in enumerate(values[:-1]):
            try:
                val = val.replace("EE", "E")
                val = nth_repl(val, ".", "-", 2) # Replace the second occurence of a period
                if val == "": # Didn't detect anything
                    warns.append((j, i))
                    continue
                if val[-1] == "E": # The number doesn't have an exponent. (Result of line 115)
                    values[j] = val[:-1]
                    continue
                if "E" in val and not val.split("E")[1][0] in ["-", "+"]: # If there is no plus or minus after the E
                    val = val.replace("E", "E-")
                if "-" in val and not val.split("-")[0][-1] == "E": # If there is no E before the minus
                    val = val.replace("-", "E-")   
                    val = str(math.floor(float(val.split("E")[0])*(10**decimals[i]))/(10.0**decimals[i])) + "E" + val.split("E")[1] # Fix the number of decimals again
                values[j] = val
            except:
                warns.append((j, i))

        # Try to convert all to floats
        for j, val in enumerate(values[:-1]):
            try:
                values[j] = float(val)
            except:
                warns.append((j, i))

    lines.append(values)
    i += 1    

# check if all numbers are ascending or descending
for i in range(len(lines[1])):
    # only check floats
    ls = list([l[i] for l in filter(lambda l : type(l[i]) is float, lines)])
    if len(ls) < 2:
        continue
    asc = (ls[0] < ls[-1])
    col = [l[i] for l in lines]
    for j in range(len(col)-1):
        #print(type(lines[j+1][i]), ls[j+1][i], ls[j][i], (ls[j+1][i] <= ls[j][i]) == asc, asc)
        if type(col[j+1]) is float and type(col[j]) is float and (col[j+1] <= col[j]) == asc:
            warns.append((i, j))

# Save everything to excel
wb = Workbook()
ws1 = wb.active
ws1.title = "Data"

for line in lines:
    ws1.append(line)

letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

orangefill = PatternFill("solid", fgColor="FF7F00")
for att in attentions:
    cell = ws1[letters[att[0]]+str(att[1]+1)]
    #print(cell)
    cell.fill = orangefill

redfill = PatternFill("solid", fgColor="FF0000")
for warn in warns:
    cell = ws1[letters[warn[0]]+str(warn[1]+1)]
    #print(cell)
    cell.fill = redfill

wb.save("p33_h1_silver.xlsx")
